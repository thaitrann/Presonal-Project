from email.mime.image import MIMEImage
import sys
import os
sys.path.insert(0, r'C:\Users\thomas.thai\Downloads\automail\automail_sale')
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
from email.mime.application import MIMEApplication
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from transform.create_folder_by_id import create_folder
from mail.account import *

def send_mail(number_, name_, mail_, mail_cc, string_date, df_pivot_merged, max_date, sum_quantity):
    start_time = time.time()
    
    #create object mail
    email_recipient_name = name_[0]
    recipient_number = number_[0]
    email_recipient = mail_[0]
    email_cc = mail_cc[0] 
    all_adrs = email_cc.split(';')
    all_adrs.append(email_recipient)

    msg = MIMEMultipart()
    msg['From'] = email_sender
    msg['To'] = email_recipient
    msg['Cc'] = email_cc

    msg['Subject'] = 'Gửi báo cáo sản lượng bán hàng định kỳ_MNS Daily sales volume report'

    # body mail
    html_content = f"""
    <html>
    <head>
    <style>
        .bold-text {{
            font-weight: bold;
            font-size: 18px;
            color: #5e94c6;
        }}
        .blue-text {{
            color: #5e94c6;
        }}
    </style>
    </head>
    <body>
        <p class="bold-text">Dear anh/chị {email_recipient_name},</p>
        <p class="blue-text">Đính kèm là báo cáo sản lượng tháng theo khu vực phụ trách. Ngày cập nhật dữ liệu: {max_date}.</p>
        <p class="blue-text">Trong trường hợp có bất cứ thắc mắc nào về sản lượng bán hàng, xin liên hệ em với thông tin sau: Thomas.Thai - 0789129691- thomas.thai@deheus.com</p>
    </body>
    </html>
    """

    # html_files = r'C:\Users\thomas.thai\Downloads\Chữ ký\chữ ký.htm'
    # with open(html_files, "r") as file:
    #     signature_html_content = file.read()

    html_files = r'C:\Users\thomas.thai\Downloads\Chữ ký\Signature.html'
    with open(html_files, "r") as file:
        signature_html_content = file.read()

    full_html = html_content + signature_html_content
    #Tạo phần MIMEText từ nội dung HTML
    html_part = MIMEText(full_html, "html")
    msg.attach(html_part)

    create_folder(email_recipient_name, recipient_number)
    
    filename_data = '{} - {} - {} - Bao cao san luong hang ngay'.format(string_date, email_recipient_name, recipient_number)
    location_data = r'C:\Users\thomas.thai\Downloads\automail\Data sending\Data gửi sale\{} - {}\{}.xlsx'.format(email_recipient_name, recipient_number, filename_data)
    
    #create obj workbook
    workbook = Workbook()
    worksheet = workbook.active

    # add value to cell J1
    worksheet["J1"] = sum_quantity

    # Format J1
    cell_J1 = worksheet["J1"]
    cell_J1.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    cell_J1.font = Font(color="FF0000", bold=True)
    cell_J1.number_format = "#,##0"

    for r in dataframe_to_rows(df_pivot_merged, index=False, header=True):
        worksheet.append(r)

    fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    font = Font(bold=True)

    #format first cell
    for cell in worksheet[2]:
        cell.fill = fill
        cell.font = font

    #scale column
    column_widths = []
    for column in df_pivot_merged:
        max_length = max(df_pivot_merged[column].astype(str).map(len).max(), len(column))
        adjusted_width = (max_length + 2) * 1.2
        column_widths.append(adjusted_width)

    for i, width in enumerate(column_widths):
        worksheet.column_dimensions[worksheet.cell(row=1, column=i+1).column_letter].width = width

    #create border
    border = Border(left=Side(border_style="thin"),
                    right=Side(border_style="thin"),
                    top=Side(border_style="thin"),
                    bottom=Side(border_style="thin"))

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border

    #save data
    workbook.save(location_data)

    #attach file
    with open(location_data, 'rb') as file:
        file_content = file.read()
    attachment = MIMEApplication(file_content)

    # Thiết lập tên file đính kèm
    attachment.add_header('Content-Disposition', 'attachment', filename='{}.xlsx'.format(filename_data))
    msg.attach(attachment)

    #send mail
    try:
        server = smtplib.SMTP(host = 'smtp.office365.com', port = 587)
        server.starttls()
        server.login(email_sender, password)
        text = msg.as_string()
        server.sendmail(email_sender, all_adrs, text)
        server.quit()
        print('Email sent successfully: {}!'.format(email_recipient_name))
    except Exception as e:
        print('Error sending email:', str(e))

    end_time = time.time()
    print("---------- DONE SEND MAIL! ----------")
    execution_time = end_time - start_time
    print("Execution time: ", execution_time)
    print("-------------------------------------")

